#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BOM 格式转换与校验工具

功能说明：
- 读取来源 CSV 文件
- 校验必要列是否存在
- 分类中文映射（按 Category 顶级分类）
- 校验 描述 与 Value 的模糊包含（忽略大小写与空白）
- 按分类和位号排序
- 导出格式化的 XLSX 文件与质量报告

使用方法：
    基础用法：
        python bom_transform.py -i <输入CSV文件> -o <输出目录>

    完整示例：
        python bom_transform.py --input "example/PV3.csv" --output-dir out --project-name "PV3"

    参数说明：
        -i, --input       [必需] 输入 CSV 文件路径
        -o, --output-dir  [可选] 输出目录，默认为当前目录
        -p, --project-name [可选] 项目名称，用于输出文件命名，默认使用输入文件名
        -m, --mapping     [可选] 映射配置文件路径（mapping.yml），默认使用内置配置
        -e, --encoding    [可选] 强制指定输入文件编码，默认自动检测

输出说明：
    1. XLSX 格式的 BOM 文件（格式化、带边框、分类合并居中）
       - 文件名格式：{项目名}_BOM_{日期}.xlsx
       - 第一列：分类（相同分类自动合并居中）
       - 位号列：自动换行，固定宽度 31.1
       - 数量列：居中显示，固定宽度 13.13
       - 其他列：自动调整宽度
       - 字体：等线，细边框

    2. TXT 格式的质量报告
       - 文件名格式：{项目名}_BOM_Report_{日期}.txt
       - 包含问题统计和详细错误信息

配置文件（mapping.yml）：
    可选的 YAML 配置文件，用于自定义：
    - required_columns: 必需的列名列表
    - output_columns: 输出列名映射
    - category_map: 分类中文映射字典
    - unknown_label: 未知分类标签
    - options: 其他选项（排序、报告格式等）

质量检查项：
    - MissingColumn: 缺失必需列
    - EmptyDescription: 描述为空
    - ValueNotInDescription: Value 未包含在描述中
    - MissingCategory: 分类为空
    - InvalidQty: 数量格式错误
"""
from __future__ import annotations
import argparse
import os
import sys
import io
import re
import datetime as dt
from typing import Dict, List, Tuple, Optional

import pandas as pd

try:
    import chardet  # type: ignore
except Exception:  # pragma: no cover
    chardet = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

try:
    import yaml  # type: ignore
except Exception:
    yaml = None

DEFAULT_MAPPING = {
    "required_columns": [
        "描述",
        "Reference",
        "Qty",
        "Value",
        "Footprint",
        "Category",
        "Part-DB IPN",
        "lcsc#",
        "manf",
        "manf#",
    ],
    "output_columns": {
        "": "Category_CN_Display",
        "规格和型号": "描述",
        "位号": "Reference",
        "数量": "Qty",
        "料号": "Part-DB IPN",
        "制造商": "manf",
        "制造商料号": "manf#",
        "LCSC 料号": "lcsc#",
    },
    "category_map": {
        "Audio": "音频器件",
        "Capacitor": "电容",
        "Connector": "连接器",
        "Crystal": "晶振",
        "Diode": "二极管",
        "Fuse": "保险丝",
        "IC": "芯片",
        "Inductor": "电感",
        "Non-BOM": "非物料",
        "PCB": "PCB板",
        "Resistor": "电阻",
        "Switch": "开关",
        "Transistor": "晶体管",
    },
    "unknown_label": "Unknown",
    "issue_type_cn": {
        "MissingColumn": "缺失必需列",
        "EmptyDescription": "描述为空",
        "ValueNotInDescription": "Value未包含在描述中",
        "MissingCategory": "分类为空",
        "InvalidQty": "数量格式错误",
    },
    "options": {
        "ignore_case_whitespace_for_value_match": True,
        "generate_report": True,
        "report_format": "txt",
        "export_format": "xlsx",
        "block_on_missing_columns": False,
        "sort_by": ["Category_CN", "Reference"],
    },
}

ISSUE_MISSING_COLUMN = "MissingColumn"
ISSUE_EMPTY_DESC = "EmptyDescription"
ISSUE_VALUE_NOT_IN_DESC = "ValueNotInDescription"
ISSUE_MISSING_CATEGORY = "MissingCategory"
ISSUE_INVALID_QTY = "InvalidQty"


def detect_encoding(path: str) -> str:
    if chardet is None:
        return "utf-8"
    try:
        with open(path, "rb") as f:
            raw = f.read(4096)
        if raw.startswith(b"\xef\xbb\xbf"):
            return "utf-8-sig"
        det = chardet.detect(raw)
        enc = det.get("encoding") or "utf-8"
        # Windows 常见回退
        if enc.lower() in {"ascii", "None"}:
            return "utf-8"
        return enc
    except Exception:
        return "utf-8"


def load_config(mapping_path: Optional[str]) -> Dict:
    config = DEFAULT_MAPPING.copy()
    if mapping_path and os.path.isfile(mapping_path) and yaml is not None:
        try:
            with open(mapping_path, "r", encoding="utf-8") as f:
                user_cfg = yaml.safe_load(f) or {}
            # 深合并
            for k, v in user_cfg.items():
                if isinstance(v, dict) and isinstance(config.get(k), dict):
                    config[k].update(v)  # type: ignore
                else:
                    config[k] = v
        except Exception as e:
            print(f"[WARN] 读取 mapping 失败，使用默认配置: {e}")
    return config


def normalize_for_match(s: str) -> str:
    # 忽略大小写和所有空白（空格/制表/换行）
    return re.sub(r"\s+", "", s or "").upper()


def compute_category_cn(
    category: Optional[str], cat_map: Dict[str, str], unknown: str
) -> str:
    if not isinstance(category, str) or category.strip() == "":
        return unknown
    top = category.split("/", 1)[0].strip()
    return cat_map.get(top, cat_map.get(top.capitalize(), unknown))


def export_to_xlsx(df: pd.DataFrame, filepath: str) -> None:
    """导出为格式化的XLSX文件，包含自动列宽、边框和合并单元格"""
    if Workbook is None:
        raise ImportError("需要安装 openpyxl 库以支持 XLSX 导出")

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    # 写入表头
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=c_idx, value=col_name)

    # 写入数据到工作表
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 定义边框样式（细边框）
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # 定义字体（等线，不加粗）
    dengxian_font = Font(name="等线", bold=False)

    # 表头样式：等线字体，居中，细边框
    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = dengxian_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 找到位号列和数量列的索引（基于列名）
    weihao_col_idx = None
    shuliang_col_idx = None
    for idx, col_name in enumerate(df.columns, 1):
        if col_name == "位号":
            weihao_col_idx = idx
        elif col_name == "数量":
            shuliang_col_idx = idx

    # 应用边框和字体到所有数据单元格
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = dengxian_font

            # 位号列：启用换行，垂直居中
            if col == weihao_col_idx:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            # 数量列：居中显示
            elif col == shuliang_col_idx:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # 其他列：垂直居中
            else:
                cell.alignment = Alignment(vertical="center")

    # 合并第一列（分类列）相同的单元格
    # 第一列是Category_CN_Display，现在每行都包含完整的分类名称
    if len(df.columns) > 0:
        # 收集所有相同分类的连续区间
        ranges = []  # [(start_row, end_row, value), ...]
        start_row = 2
        prev_value = df.iloc[0, 0] if len(df) > 0 else None

        for idx in range(len(df)):
            current_value = df.iloc[idx, 0]
            row_num = idx + 2

            if current_value != prev_value:
                # 保存上一个区间
                ranges.append((start_row, row_num - 1, prev_value))
                start_row = row_num
                prev_value = current_value

        # 保存最后一个区间
        if len(df) > 0:
            ranges.append((start_row, len(df) + 1, prev_value))

        # 对每个区间进行合并和居中
        for start, end, value in ranges:
            if start <= end:
                if start == end:
                    # 单行：只设置对齐和字体
                    cell = ws.cell(row=start, column=1)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = dengxian_font
                else:
                    # 多行：合并单元格
                    ws.merge_cells(
                        start_row=start, start_column=1, end_row=end, end_column=1
                    )
                    merged_cell = ws.cell(row=start, column=1)
                    merged_cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    merged_cell.font = dengxian_font

    # 自动调整列宽
    for col_idx, column in enumerate(df.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        # 检查列名长度
        max_length = len(str(column))

        # 特殊处理位号列和数量列
        if column == "位号":
            # 位号列固定宽度为31.1
            ws.column_dimensions[column_letter].width = 31.1
        elif column == "数量":
            # 数量列固定宽度为13.13
            ws.column_dimensions[column_letter].width = 13.13
        else:
            # 其他列：检查每个单元格的内容长度
            for row_idx in range(len(df)):
                cell_value = str(df.iloc[row_idx, col_idx - 1])
                # 中文字符按2个字符宽度计算
                cell_length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                max_length = max(max_length, cell_length)

            # 设置列宽（加一些余量）
            adjusted_width = min(max_length + 2, 100)  # 最大宽度限制为100
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)


def transform(df: pd.DataFrame, cfg: Dict) -> Tuple[pd.DataFrame, pd.DataFrame]:
    required = cfg["required_columns"]
    out_map = cfg["output_columns"]
    cat_map = cfg.get("category_map", {})
    unknown = cfg.get("unknown_label", "Unknown")
    issue_type_cn = cfg.get("issue_type_cn", {})
    options = cfg.get("options", {})

    issues = []  # list of dicts

    # 记录缺失列
    missing_cols = [c for c in required if c not in df.columns]
    for c in missing_cols:
        issues.append(
            {
                "Reference": "",
                "描述": "",
                "Value": "",
                "IssueType": ISSUE_MISSING_COLUMN,
                "Detail": f"缺失必需列: {c}",
            }
        )

    # 对存在的列进行基本清洗（去首尾空白）
    for c in df.columns:
        if pd.api.types.is_string_dtype(df[c]):
            df[c] = df[c].astype(str).map(lambda x: x.strip())

    # 处理 Qty 尝试转 int
    if "Qty" in df.columns:

        def to_int_or_keep(x):
            if pd.isna(x):
                return x
            try:
                xi = int(float(str(x).strip()))
                return xi
            except Exception:
                issues.append(
                    {
                        "Reference": "",
                        "描述": str(x),
                        "Value": "",
                        "IssueType": ISSUE_INVALID_QTY,
                        "Detail": f"数量格式无效: {x}",
                    }
                )
                return x

        df["Qty"] = df["Qty"].map(to_int_or_keep)

    # 生成中文分类列
    if "Category" in df.columns:
        df["Category_CN"] = df["Category"].map(
            lambda s: compute_category_cn(s, cat_map, unknown)
        )

        # 提取英文顶级分类
        def extract_top_category(cat):
            if not isinstance(cat, str) or cat.strip() == "":
                return ""
            return cat.split("/", 1)[0].strip()

        df["Category_Top"] = df["Category"].map(extract_top_category)
        # 对空类别记录 issue
        mask_empty_cat = df["Category"].isna() | (
            df["Category"].astype(str).str.strip() == ""
        )
        for idx in df[mask_empty_cat].index:
            issues.append(
                {
                    "Reference": (
                        df.at[idx, "Reference"] if "Reference" in df.columns else ""
                    ),
                    "描述": df.at[idx, "描述"] if "描述" in df.columns else "",
                    "Value": df.at[idx, "Value"] if "Value" in df.columns else "",
                    "IssueType": ISSUE_MISSING_CATEGORY,
                    "Detail": "分类字段为空",
                }
            )
    else:
        df["Category_CN"] = unknown
        df["Category_Top"] = ""

    # 描述与 Value 模糊包含校验
    if (
        options.get("ignore_case_whitespace_for_value_match", True)
        and "描述" in df.columns
        and "Value" in df.columns
    ):
        desc_norm = df["描述"].fillna("").map(normalize_for_match)
        val_norm = df["Value"].fillna("").map(normalize_for_match)

        # 空描述
        for idx in df[
            df["描述"].isna() | (df["描述"].astype(str).str.strip() == "")
        ].index:
            issues.append(
                {
                    "Reference": (
                        df.at[idx, "Reference"] if "Reference" in df.columns else ""
                    ),
                    "描述": df.at[idx, "描述"] if "描述" in df.columns else "",
                    "Value": df.at[idx, "Value"] if "Value" in df.columns else "",
                    "IssueType": ISSUE_EMPTY_DESC,
                    "Detail": "描述字段为空",
                }
            )

        # Value 未包含于 描述（逐行比较，忽略大小写与空白）
        # pandas 对 Series vs Series 的 str.contains 支持有限，这里用逐行判断以避免类型问题
        contains_list = []
        for a, b in zip(desc_norm.tolist(), val_norm.tolist()):
            if not b:  # 空 Value 视为不检测
                contains_list.append(True)
            else:
                contains_list.append(b in a)
        contains_mask = pd.Series(contains_list, index=df.index)
        mismatch = (val_norm != "") & (~contains_mask)
        for idx in df[mismatch].index:
            issues.append(
                {
                    "Reference": (
                        df.at[idx, "Reference"] if "Reference" in df.columns else ""
                    ),
                    "描述": df.at[idx, "描述"] if "描述" in df.columns else "",
                    "Value": df.at[idx, "Value"] if "Value" in df.columns else "",
                    "IssueType": ISSUE_VALUE_NOT_IN_DESC,
                    "Detail": "Value值未包含在描述中（忽略大小写和空格的比对）",
                }
            )

    # 排序
    sort_keys = options.get("sort_by", ["Category_CN", "Reference"])
    for k in sort_keys:
        if k not in df.columns:
            # 允许缺失排序键，跳过
            sort_keys = [x for x in sort_keys if x in df.columns]
            break
    if sort_keys:
        df = df.sort_values(
            by=sort_keys, kind="stable", na_position="last"
        ).reset_index(drop=True)

    # 创建用于显示的分类列 - 必须在排序后执行
    # 注意：不再将相同分类设为空字符串，而是保持原值，通过Excel合并单元格来实现视觉效果
    df["Category_CN_Display"] = df["Category_CN"].copy()

    # 构造输出
    out_cols_order = list(cfg["output_columns"].keys())
    out_df = pd.DataFrame()
    for out_col, src_col in cfg["output_columns"].items():
        out_df[out_col] = df[src_col] if src_col in df.columns else ""

    # 报告 DataFrame
    report_df = pd.DataFrame(
        issues, columns=["Reference", "描述", "Value", "IssueType", "Detail"]
    )

    return out_df, report_df


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="BOM CSV 转换与校验")
    parser.add_argument("--input", "-i", required=True, help="输入 CSV 文件路径")
    parser.add_argument("--output-dir", "-o", default=".", help="输出目录")
    parser.add_argument(
        "--project-name", "-p", default=None, help="项目名（用于命名导出文件）"
    )
    parser.add_argument(
        "--mapping", "-m", default=None, help="映射配置 mapping.yml（可选）"
    )
    parser.add_argument(
        "--encoding", "-e", default=None, help="强制指定输入编码（可选）"
    )
    parser.add_argument(
        "--quiet",
        "-q",
        action="store_true",
        help="静默模式，不输出详细的错误报告到终端",
    )
    args = parser.parse_args(argv)

    cfg = load_config(args.mapping)

    in_path = args.input
    if not os.path.isfile(in_path):
        print(f"[ERROR] 输入文件不存在: {in_path}")
        return 2

    enc = args.encoding or detect_encoding(in_path)

    try:
        df = pd.read_csv(in_path, encoding=enc)
    except UnicodeDecodeError:
        # 退回尝试 utf-8-sig / gbk
        for alt in ("utf-8-sig", "gbk", "utf-8"):
            try:
                df = pd.read_csv(in_path, encoding=alt)
                enc = alt
                break
            except Exception:
                continue
        else:
            print(f"[ERROR] 无法解码文件: {in_path}")
            return 3

    out_df, report_df = transform(df, cfg)

    # 输出目录与文件名
    os.makedirs(args.output_dir, exist_ok=True)
    base_project = args.project_name or os.path.splitext(os.path.basename(in_path))[0]
    date_tag = dt.datetime.now().strftime("%Y%m%d")

    export_format = cfg.get("options", {}).get("export_format", "csv").lower()

    if export_format == "xlsx":
        out_file = os.path.join(args.output_dir, f"{base_project}_BOM_{date_tag}.xlsx")
        export_to_xlsx(out_df, out_file)
    else:
        out_file = os.path.join(args.output_dir, f"{base_project}_BOM_{date_tag}.csv")
        out_df.to_csv(out_file, index=False, encoding="utf-8-sig")

    rep_file = os.path.join(
        args.output_dir, f"{base_project}_BOM_Report_{date_tag}.txt"
    )

    if cfg.get("options", {}).get("generate_report", True):
        # 写入TXT格式报告
        with open(rep_file, "w", encoding="utf-8") as f:
            f.write("BOM 转换质量报告\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"项目名称: {base_project}\n")
            f.write(f"生成时间: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"输入文件: {in_path}\n")
            f.write(f"输入编码: {enc}\n")
            f.write(f"总行数: {len(df)}\n\n")

            if not report_df.empty:
                summary = (
                    report_df.groupby("IssueType").size().sort_values(ascending=False)
                )
                f.write("问题统计\n")
                f.write("-" * 80 + "\n")
                for t, n in summary.items():
                    f.write(f"  {t}: {n}\n")
                f.write("\n")

                f.write("错误详情\n")
                f.write("=" * 80 + "\n\n")
                issue_type_cn_map = cfg.get("issue_type_cn", {})
                for idx, row in report_df.iterrows():
                    issue_type_cn = issue_type_cn_map.get(
                        row["IssueType"], row["IssueType"]
                    )
                    f.write(f"[{issue_type_cn}] 位号: {row['Reference']}\n")
                    f.write(f"  详情: {row['Detail']}\n")
                    if row["描述"]:
                        f.write(f"  描述: {row['描述']}\n")
                    if row["Value"]:
                        f.write(f"  Value: {row['Value']}\n")
                    f.write("\n")
            else:
                f.write("未发现质量问题。\n")

    # 控制台汇总
    print(f"输入编码: {enc}")
    print(f"输出文件: {out_file}")
    if cfg.get("options", {}).get("generate_report", True):
        print(f"报告文件: {rep_file}")
        if not report_df.empty:
            summary = report_df.groupby("IssueType").size().sort_values(ascending=False)
            print("\n========== 问题统计 ==========")
            for t, n in summary.items():
                print(f"  - {t}: {n}")

            # 仅在非静默模式下输出所有错误详情到终端
            if not args.quiet:
                print("\n========== 错误报告详情 ==========")
                issue_type_cn_map = cfg.get("issue_type_cn", {})
                for idx, row in report_df.iterrows():
                    issue_type_cn = issue_type_cn_map.get(
                        row["IssueType"], row["IssueType"]
                    )
                    print(f"[{issue_type_cn}] {row['Reference']}: {row['Detail']}")
                    if row["描述"]:
                        print(f"  描述: {row['描述']}")
                    if row["Value"]:
                        print(f"  Value: {row['Value']}")
                    print()
            else:
                print("(详细错误报告已保存到文件，使用 -q 参数已隐藏终端输出)")
        else:
            print("问题统计：无不合规记录")

    return 0


if __name__ == "__main__":
    sys.exit(main())
